"""Delete a batch of Samsara addresses by ID.

The script accepts a CSV or Excel file containing a column named ``ID``. Each value in
that column is passed to :meth:`SamsaraClient.delete_address`.

Example usage::

    export SAMSARA_BEARER_TOKEN=your_token
    python -m encompass_to_samsara.scripts.delete_addresses ids.csv

"""

from __future__ import annotations

import argparse
import csv
import os
import sys
from pathlib import Path

from openpyxl import load_workbook

from ..samsara_client import SamsaraClient


def _load_ids(path: Path) -> list[str]:
    """Return address IDs from ``path``.

    ``path`` may point to a ``.csv`` or ``.xlsx`` file. Files must contain a column
    header ``ID``.
    """

    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames is None or "ID" not in reader.fieldnames:
                msg = "CSV must contain an 'ID' column"
                raise ValueError(msg)
            return [row["ID"].strip() for row in reader if row.get("ID")]

    if path.suffix.lower() == ".xlsx":
        wb = load_workbook(filename=path, read_only=True)
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1))
        headers = [c.value for c in header_row]
        try:
            idx = headers.index("ID")
        except ValueError as exc:  # pragma: no cover - path exercised in tests
            raise ValueError("Excel file must contain an 'ID' column") from exc
        ids: list[str] = []
        for row in ws.iter_rows(min_row=2):
            val = row[idx].value
            if val is not None and str(val).strip():
                ids.append(str(val).strip())
        return ids

    msg = "Unsupported file type; use .csv or .xlsx"
    raise ValueError(msg)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Delete Samsara addresses specified in a CSV or Excel file"
    )
    parser.add_argument("input", help="Path to CSV or Excel file with an 'ID' column")
    args = parser.parse_args(argv)

    ids = _load_ids(Path(args.input))
    client = SamsaraClient(api_token=os.environ["SAMSARA_BEARER_TOKEN"])

    for addr_id in ids:
        try:
            client.delete_address(addr_id)
            print(f"Deleted {addr_id}")
        except Exception as exc:  # pragma: no cover - network errors
            print(f"Failed to delete {addr_id}: {exc}", file=sys.stderr)


if __name__ == "__main__":
    main()

